#RPDataProcessing.py Python app to process Reck Peterson Lab Data in XLSX format
#Author: James Holcomb
#Last Revised: July 19th, 2020

#Library imports
import os                                   #imports functions for interacting with O/S - allows Command Line 
import re                                   #imports regular expression operations
import openpyxl                             #imports Python Library to read/write 2010 Excel files
import requests                             #imports gives ability to send HTTP/1.1 request
import csv                                  #imports functions for csv parsing
import tkinter as tk                        #imports interface to the Tk GUI toolkit
from tkinter import filedialog
import xml.etree.ElementTree as ET
import sys                                  #imports ability to allow cmd line arguments to be passed to the program
import imp                                  #imports ability to dynamically import tkinter if operating in non-headless mode


#----------------------------------------------------------------------------------
# Function to process the CSV search results retrieved by the RID value 
#----------------------------------------------------------------------------------
def process_csv_output(RIDoutput, outputFile, gene, fold, pvalue, remote, removeDups):
    previousAccession = ""
    with open(RIDoutput) as csv_file:
        csv_reader = csv.reader(csv_file, dialect='excel')
        line_count = 0
        hit_count = 1
        for row in csv_reader:
            if (remote == 1 and line_count == 0) or ((removeDups[0] == "Y" or removeDups[0] == "y") and row[6] == previousAccession):
                # skip the column name row
                line_count += 1
            else:
                if len(row) > 5:   # avoid garbage lines with tabs and such
                    qCov = int(row[3].rstrip("%"))
                    if qCov < 20:
                        print("Skipping " + row[0] + " result as QCov is:" + row[3])
                        line_count += 1
                    else:
                        col_count = 0
                        if remote == 1:                 # Stripping quotes on remote searches so that the accession hyperlinks work.
                            for x in row:   
                                row[col_count] = x.replace('"', '""')   # replace stripped double quotes
                                if (col_count == 0 or col_count == 6):
                                    row[col_count] = '"' + row[col_count] + '"'   # add quotes to 1st and last string fields
                                col_count += 1

                        # Could add additional filtering to the RID output if desired
                        # Output the input data
                        outputFile.write(gene + "," + fold + "," + pvalue)
                        #output hit number
                        outputFile.write(f',{hit_count}')
                        hit_count += 1
                        #output the search result data fields we want
                        outputFile.write(f',{row[0]},{",".join(row[3:])}')
                        # outputFile.write(f',"{row[7]}"')
                        outputFile.write("\n")
                        if removeDups[0] == "Y" or removeDups[0] == "y":
                            previousAccession = row[6]
                        line_count += 1

        if remote == 1 and line_count < 2:   # then we did not have any results
             # Output the input data
            outputFile.write(gene + ", " + fold + ", " + pvalue + ", ")
            outputFile.write("0, 0, 0, 0, No matches found \n")


#----------------------------------------------------------------------------------
# Function to parse and use the RID value to get desired results file
#----------------------------------------------------------------------------------
def parse_RID(outputFile):
    #read blastp results ID (RID) from output file
    with open(outputFile, "r") as blastp:
        blpl = blastp.readlines()

        for x in blpl:
        
            #check if blank line
            if not x.isspace():
                                    
                #check for RID
                if "RID:" in x:
                    RID = x[x.find(ridTxt) + ridTxtLength:]
                    RID = RID.strip()
                    
                    # How to get the results file from the RID
                    # Extract the RID from the original blastp output then replace the RID in the following command.
                    Output1 = "https://blast.ncbi.nlm.nih.gov/Blast.cgi?RESULTS_FILE=on&RID="
                    Output2 = "&FORMAT_TYPE=CSV&DESCRIPTIONS=100&FORMAT_OBJECT=Alignment&QUERY_INDEX=0&DOWNLOAD_TEMPL=Results&CMD=Get&RID="
                    Output3 = "&ALIGNMENT_VIEW=Pairwise&QUERY_INDEX=0&CONFIG_DESCR=2,3,4,5,6,7,8"
                    RIDURL = Output1 + RID + Output2 + RID + Output3
                    print(RIDURL)
                    RIDOutputResponse = requests.get(RIDURL, allow_redirects=True)

#Comment out for debug ^

                    RIDoutput = outDir + RID + "_output.csv"
                    open(RIDoutput, 'wb').write(RIDOutputResponse.content)
                    
#Comment out for debug ^

                    break #RID found and processed. Leaves for loop.

        if RID == "" :
            print("RID " + ridTxt + " not found in blast output file")
            return "Failure to get RID"
        else:
            return RIDoutput

#----------------------------------------------------------------------------------
# Main Program
#----------------------------------------------------------------------------------
#Command line arguments to be passed to program
print ('Number of arguments:', len(sys.argv), 'arguments.')
print ('Argument List:', str(sys.argv))
print ("Current working directory: ", os. getcwd())
print ("Program name = ", sys.argv[0])

#Files used later in App
currentFastaFile = "currentFasta.txt"
blastpFile = "blastpResults.out"
logFileName = "logFile.out"
output_csv_name = "match_output.csv"
# Choice of remote or local database processing
remote = int(input("Enter 1 to process your file remotely with NCBI servers or enter 2 to process it using a local database: "))
remote = max(min(remote, 2),1)
if remote == 1:
    print("Processing remote.\n")
    removeDups = "N"                    #Have not currently noticed any duplicate results in remote searches.
else:
    print("Processing locally, expected db name is nr.\n")
    removeDups = str(input("Remove duplicated results? Type [Y]es or [N]o. \n"))

if len(sys.argv) > 1:    #This is what occurs if starting with an absolute file path in the command line
    wbName = sys.argv[1]
else:                   #This is what occurs if not given an absolute file path (uses Tkinter)
    root = tk.Tk()
    root.withdraw()
    wbName = filedialog.askopenfilename()

print("input name " + wbName)
wbName = os.path.normpath(wbName)
print("input name after OS normalization " + wbName)
outDir = os.path.dirname(wbName) + "\\"
#print("outdir is " + outDir)
outName = os.path.basename(wbName)
#print("outName is " + outName)
(outName,extension) = os.path.splitext(outName)
#print("outName is now " + outName)
outName = outDir + outName + "_results.csv"
print("outName is " + outName)
addHeaders = False
try:
    outputFile = open(outName, "r")
    print("The output file " + outName + " already exists! Appending to existing file!")
    outputFile.close()
except:
    addHeaders = True

outputFile = open(outName, "a")


#URL and Command Query 
uniprotURLString1 = 'https://www.uniprot.org/uniprot/?query="'
uniprotURLString2 = '"&fil=organism:\"Homo+sapiens+(Human)+[9606]\"&sort=score&columns=id&format=tab' 
fastaURLString1 = "https://www.uniprot.org/uniprot/"
fastaURLString2 = ".fasta"

#--The primary remote blastp query ---------------------------------------------------------------------------
if remote == 1:
    blastpQuery = 'cmd /c "echo "Calling Blastp " & blastp -db nr -query  "' + outDir + currentFastaFile + '" -entrez_query "Aspergillus nidulans FGSC A4"  -out "' + outDir + blastpFile + '" -remote -qcov_hsp_perc 20 -outfmt "7 sseqid" & echo "Blastp Finished""'

else:
    blastpQuery = 'cmd /c "echo "Calling Blastp " & blastp -db nr -query  "' + outDir + currentFastaFile + '" -taxids 227321 -out "' + outDir + blastpFile + '" -qcov_hsp_perc 20 -outfmt "10 ssciname score length qcovs evalue pident sacc" & echo "Blastp Finished""'
print ("new bp query is " + blastpQuery)

#Hard-coded Parameters used later in App
foldCheck = 1.6     #foldCheck and pvalueCheck are the conditional formatting for the highlighting of cells
pvalueCheck = 1.3
scoreMin = 20
eValueMax = 1

#Initialization of Parameters used later in App
outRow = 1    #set initial data row
log2Fold = 0
log10PValue = 0

#set delimiters and delimiter lengths for blastp text
ridTxt = "RID: "
ridTxtLength = len(ridTxt)

#Open and set logFile (for debugging purposes)
logFile = open(outDir+logFileName, "w")

#output column names to outputfile depending on addHeaders Boolean
if addHeaders == True:
    print("Putting header into output file")
    outputFile.write("NCBI_GENE, log2(fold), log10(pvalue), Hit Number, Description, Query Cover, E Value, % Ident, Accession \n")
#outputFile.close

#Open and set Spreadsheet up
wb = openpyxl.load_workbook(wbName)
ws_count = 0
sheetList = []
for sheet in wb:
    print("% s" %ws_count + ": " + sheet.title)
    sheetList.append(sheet.title)
    ws_count+=1
wsNumber = int(input("Please enter the number of the worksheet you want processed: "))
ws = wb[sheetList[wsNumber]]
print("Working on sheet " + ws.title)

wsOutName = ws.title + " Processed Data"
print("worksheet name " + ws.title)

#wsOut = wb[wsOutName]

#start processing rows of spreadsheet data
topValue = int(input("Enter the first row you want processed: "))
topValue = max(2, topValue)
botValue = int(input("Enter the last row you want processed: ")) + 1
for i in range(topValue, botValue): #ws.max_row):               #skip header row start with 2
    
    NCBIgene = ws.cell(row=i, column=1).value   #extract NCBI Gene information

    #Need to check if gene is empty or whitespace;
    if (not (str(NCBIgene).isspace() or NCBIgene is None)): 

        #extract log2(fold) and log10(pvalue) data
        try:
            log2Fold_txt = str(ws.cell(row=i, column=2).value) 
            log2Fold = float(log2Fold_txt)
        except ValueError:
            logFile.write("ERROR - Expected log2Fold Value\n")
            logFile.write("x: "+ log2Fold_txt + "\n")
            break
        
        try:
            log10PValue_txt = str(ws.cell(row=i, column=3).value)   
            log10PValue = float(log10PValue_txt)
        except ValueError:
            logFile.write("ERROR - Expected log10PValue Value\n")
            logFile.write("x: "+ log10PValue_txt + "\n")
            break
        
        currentProtein = "Skipped or No protein found"
        print("Row " + str(i) + " Gene name is: " + NCBIgene)

        #Check if log2(fold) and log10(pvalue) data falls within bounds to check for potential match           
        if (log2Fold > foldCheck and log10PValue > pvalueCheck):     #Highlighting is conditional so checking value for Columns B and C
         
            #generate uniprot URL, and extract protein
            uniprotURL = uniprotURLString1 + NCBIgene + uniprotURLString2
            #print("uniprotURL is: " + uniprotURL)
            uniprotResponse = requests.get(uniprotURL)
            #print ("Uniprot response is " + uniprotResponse.text)
            currentProtein = uniprotResponse.text.split('\n')[1]
            print ("Current protein: " + currentProtein)
            
            #generate URL to get FASTA, and write to the fasta file used in the query
            fastaURL = fastaURLString1 + currentProtein + fastaURLString2
            fastaResponse = requests.get(fastaURL)  
            fasta = fastaResponse.text
            logFile.write("Processing Protein: "+ currentProtein +"\n")
            print("Processing Uniprot Protein: "+ currentProtein +"\n")
            print("Fasta file path: " + outDir+currentFastaFile)
            with open(outDir+currentFastaFile, "w") as fastafile:
                fastafile.write(fasta)

            #run Command Line to query blastp
            logFile.close()    #closing because os.system command messes up open file
            os.system(blastpQuery)

#Comment out for debug ^

            logFile = open(outDir+logFileName, "a")    # reopening log file

            if remote == 1:
                RIDoutput = parse_RID(outDir+blastpFile)
                
                # Then parse the RID CSV results file and write the desired results to the output file
                process_csv_output(RIDoutput, outputFile, NCBIgene, log2Fold_txt, log10PValue_txt, remote, removeDups)
            else:
                # Then parse the blastp results file and write the desired results to the output file
                process_csv_output(outDir+blastpFile, outputFile, NCBIgene, log2Fold_txt, log10PValue_txt, remote, removeDups)
 
            #finished processing sequence                     
            logFile.write("Done Processing Sequence\n")
        else:
            logFile.write(NCBIgene + " did not meet the minimum requirements and was skipped.")

        logFile.write("Done Processing Protein: "+ currentProtein +"\n")
        print("Done Processing Protein: "+ currentProtein +"\n")

    else:
        print("Skipping row " + str(i) + " because it is blank \n")
        
    # wb.save(filename = wbOutName)   #save output spreadsheet data after processing each row
logFile.close() 


